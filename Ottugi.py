# -*- coding:utf-8 -*-

import sys
import openpyxl
import os
import time

dirpath = "D:\\Ottugi"

# 열 너비 자동 맞춤
# columns is passed by list and element of columns means column index in worksheet.
# if columns = [1, 3, 4] then, 1st, 3rd, 4th columns are applied autofit column.
# margin is additional space of autofit column.

def AutoFitColumnSize(worksheet, columns=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns is None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

def find_last_row(ws, side_col, max_row):
    while True:
        last_row = str(side_col) + str(max_row)

        # print(ws_master[last_row])
        if ws[last_row].value is None:
            max_row -= 1
        else:
            break

    return max_row


def copy_to_paste_column():
    for file in file_list:  # 파일 당 하나씩
        wb_master = openpyxl.load_workbook(result_file_name)
        ws_master = wb_master.active
        wb_slave = openpyxl.load_workbook(file)
        ws_slave = wb_slave.active  # wb["sheetname"]

        for column_info in column_match_list:  # 복붙 할 컬럼 당 하나씩 (알파벳 다 끝나고 마스터 갱신 )
            copy_side_col = column_info[0]
            paste_side_col = column_info[1]

            master_max_row = find_last_row(ws_master, paste_side_col, ws_master.max_row) + 1
            slave_max_row = find_last_row(ws_slave, copy_side_col, ws_slave.max_row)

            print(f'[{file}] ==> master : {master_max_row}')
            print(f'[{file}] ==> slave : {slave_max_row}')

            cell_range = str(copy_side_col) + str(start_row) + ":" + str(copy_side_col) + str(slave_max_row)
            print(cell_range)

            copied_column_list = []
            # 복사
            for row in ws_slave[cell_range]:
                # print(row[0])
                copied_column_list.append(row[0])

            for i, data in enumerate(copied_column_list):
                paste_cell = str(paste_side_col) + str(master_max_row + i)
                print(paste_cell)
                ws_master[paste_cell] = data.value

        print("save")
        wb_master.save(result_file_name)
        wb_master.close()
        wb_slave.close()

def input_column_info():
    while True:
        column_match_list = []
        print("== 빈 값이 입력되면 진행됩니다 ==")
        print("복사할 행 / 붙여넣을 행 순서대로 입력")
        print("예시 : B A")

        while True:
            column_info = list(map(str, input().split()))

            if not column_info:
                break

            column_match_list.append(column_info)

        print("=========== 컬럼 정보 ===========")
        for column_info in column_match_list:
            print(f'[{column_info[0]}] ==> [{column_info[1]}]')

        ANSWER = input("\n\nContinue ? (Y/N) ")
        if ANSWER == "n" or ANSWER == "N":
            continue

        elif ANSWER != "y" or ANSWER != "Y":
            return column_match_list
        else:
            continue


def print_filelist(flist):
    for num, file in enumerate(flist):
        print(f'\t\t[{num + 1}] ==> {file}')

    # Ask Continue
    while True:
        ANSWER = input("\n\nContinue ? (Y/N) ")

        if ANSWER == "n" or ANSWER == "N":
            sys.exit(0)
        elif ANSWER != "y" or ANSWER != "Y":
            break
        else:
            continue


def make_filelist():
    print("=============== Directory Check ================")
    if os.path.isdir(dirpath) is False:
        print(f'{dirpath} doesn\'t exist. Check it')
        sys.exit(0)
    else:
        print(f"[{dirpath}] OK")

    flist = [os.path.join(dirpath, f) for f in os.listdir(dirpath) if os.path.isfile(os.path.join(dirpath, f))]

    return flist


def find_result_filename():
    temp = input("Saved result file: ")
    result_name = os.path.join(dirpath, temp)

    return result_name


if __name__ == '__main__':
    # 0. 결과파일을 저장할 파일이름을 입력받음
    result_file_name = find_result_filename()

    # 1. 해당 디렉토리에 있는 파일리스트 출력
    file_list = make_filelist()
    file_list.remove(result_file_name)  # 리스트 중에서 결과를 저장하는 xlsx 파일은 제외

    # 2. 파일 리스트 내용 출력
    print_filelist(file_list)

    # 4. 복사 / 붙여넣기 할 컬럼 입력받음
    column_match_list = input_column_info()

    # 5. 복사 시작 할 행 번호 입력받음
    start_row = int(input("복사 시작할 행 번호 입력 : "))

    # 6. 엑셀 복사 붙여넣기
    copy_to_paste_column()

