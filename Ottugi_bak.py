# -*- coding:utf-8 -*-

import sys
import openpyxl
import os
import time

dirpath = "D:\\Ottugi"

# 열 너비 자동 맞춤
# columns is passed by list and element of columns means column index in worksheet.
# if columns = [1, 3, 4] then, 1st, 3rd, 4th columns are applied autofit column.
# margin is additional space of autofit column.

def AutoFitColumnSize(worksheet, columns=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns is None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet


if __name__ == '__main__':
    # 0. 결과파일을 저장할 파일 입력
    temp = input("Saved result file: ")
    result_file_name = os.path.join(dirpath, temp)

    # 1. 파일 리스트 출력
    print("=============== Directory Check ================")
    if os.path.isdir(dirpath) is False:
        print(f'{dirpath} doesn\'t exist. Check it')
        sys.exit(0)
    else:
        print(f"[{dirpath}] OK")

    file_list = [os.path.join(dirpath, f) for f in os.listdir(dirpath) if os.path.isfile(os.path.join(dirpath, f))]

    file_list.remove(result_file_name)

    for num, file in enumerate(file_list):
        print(f'\t\t[{num + 1}] ==> {file}')

    # while True:
    ANSWER = input("\n\nContinue ? (Y/N) ")
    if ANSWER == "n" or ANSWER == "N":
        sys.exit(0)
    elif ANSWER != "y" or ANSWER != "Y":
        pass  # break
    else:
        pass
    # else:
    #     # continue    : 다시 입력받도록

    # 1.5 정보 입력받기
    '''

    '''
    while True:
        column_match_list = []
        print("== 빈 값이 입력되면 진행됩니다 ==")
        print("복사할 행 / 붙여넣을 행 순서대로 입력")
        print("예시 : B A")

        while True:
            column_info = list(map(str, input().split()))
            # upper case 해주기

            if not column_info:
                break

            column_match_list.append(column_info)

        print("=========== 컬럼 정보 ===========")
        for column_info in column_match_list:
            print(f'[{column_info[0]}] ==> [{column_info[1]}]')

        ANSWER = input("\n\nContinue ? (Y/N) ")
        if ANSWER == "n" or ANSWER == "N":
            continue

        elif ANSWER != "y" or ANSWER != "Y":
            break
        else:
            continue

    start_row = int(input("복사 시작할 행 번호 입력 : "))

    # # 행에 대한 딕셔너리 생성
    # copied_dict = {}
    # copied_dict.setdefault()

    # 2. 엑셀 열기
    for file in file_list:  # 파일 당 하나씩
        wb_master = openpyxl.load_workbook(result_file_name)
        ws_master = wb_master.active
        wb_slave = openpyxl.load_workbook(file)
        ws_slave = wb_slave.active  # wb["sheetname"]

        for column_info in column_match_list:  # 복붙 할 컬럼 당 하나씩 (알파벳 다 끝나고 마스터 갱신 )
            copy_side_col = column_info[0]
            paste_side_col = column_info[1]

            print(f'copy_side_col : {copy_side_col}')
            print(f'paste_side_col : {paste_side_col}')

            # 붙여넣을 셀 분석
            # temp = ws_master[paste_side_col]
            # print(temp)
            # master_max_row = len(temp)

            # # 복사하려는 셀 열의 최대값 찾기
            # slave_max_row = len(ws_slave[copy_side_col])
            # # slave_max_row = len(ws_slave[temp])

            master_max_row = ws_master.max_row
            slave_max_row = ws_slave.max_row

            while True:
                last_row = str(paste_side_col) + str(master_max_row)

                # print(ws_master[last_row])
                if ws_master[last_row].value is None:
                    master_max_row -= 1
                else:
                    break

            master_max_row += 1

            while True:
                last_row = str(copy_side_col) + str(slave_max_row)

                if ws_slave[last_row].value is None:
                    slave_max_row -= 1
                else:
                    break

            print(f'[{file}] ==> master : {master_max_row}')
            print(f'[{file}] ==> slave : {slave_max_row}')

            cell_range = str(copy_side_col) + str(start_row) + ":" + str(copy_side_col) + str(slave_max_row)
            print(cell_range)

            copied_column_list = []
            # 복사
            for row in ws_slave[cell_range]:
                # print(row[0])
                copied_column_list.append(row[0])

            for i, data in enumerate(copied_column_list):
                paste_cell = str(paste_side_col) + str(master_max_row + i)
                print(paste_cell)
                ws_master[paste_cell] = data.value

        print("save")
        wb_master.save(result_file_name)
        wb_master.close()
        wb_slave.close()

    # wb_master = openpyxl.load_workbook(result_file_name)
    # ws_master = wb_master.active

    # temp = ws_master.max_row
    # print(temp)
    # for i in ws_master.iter_rows(min_row=0):
    #     print("len is : ", len(i))

    # max_row_for_c = max((c.row for c in ws['C'] if c.value is not None))

    # for i in range(1, m_row + 1):
    # # for j in range(1, m_col + 1):
    #     c = ws_slave.cell(row = i, column= 4)   # 열 고정, 행 변화
    #     ws_master.cell(row = i, column = 4).value = c.value

    # wb_master.save(str(result_file_name))
